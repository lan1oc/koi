import json
import os
import re
from collections import Counter

# 导入分类逻辑（来自 group_folders.py）
COMPANY_KEYWORDS = [
    "公司", "集团", "股份",
    "有限责任公司", "有限公司",
    "制造厂", "工厂", "厂",
    "店", "中心", "研究所", "研究院", "医院", "学校",
    "商行", "事务所", "合作社", "农场", "工作室",
    "局", "厅", "处", "署", "队", "站", "网",
    "超市", "经营部", "便利店", "饭店", "酒店", "宾馆", "旅馆",
    "网吧", "俱乐部", "棋牌", "会所", "KTV", "吧",
    "委员会", "协会", "党支部", "联合会",
    "办事处",
    "小学", "中学", "初中", "高中", "大学", "幼儿园", "托儿所"
]

def is_company_line(line: str) -> bool:
    s = line.strip()
    if not s:
        return False
    return any(k in s for k in COMPANY_KEYWORDS)

def normalize_company(name: str):
    s = name.strip()
    # 1. 清理常见前缀
    s = re.sub(r'^[（(【]专项[）)】]', '', s)
    s = re.sub(r'^[（(【\[][^）)】\]]+[）)】\]]', '', s)
    s = re.sub(r'^关于', '', s)
    s = re.sub(r'^通报[：:]', '', s)
    s = s.strip()
    
    # 2. 提取“所属”之前的部分
    if "所属" in s:
        s = s.split("所属")[0].strip()
    
    # 3. 定义后缀优先级
    # 强后缀：通常是企业或机构的正式结尾
    strong_suffixes = [
        "股份有限公司", "有限责任公司", "有限公司", "责任有限公司", 
        "集团公司", "集团", "公司", "制造厂", "工厂", "厂",
        "中心", "研究所", "研究院", "医院", "学校", "幼儿园", "托儿所",
        "商行", "事务所", "合作社", "农场", "经营部", "工作室",
        "委员会", "协会", "党支部", "联合会", "超市", "便利店",
        "饭店", "酒店", "宾馆", "旅馆"
    ]
    
    # 弱后缀：可能是公司名一部分，也可能是描述词（如“网站”、“僵尸网”）
    weak_suffixes = [
        "局", "厅", "处", "署", "队", "站", "网", "店", "吧", "KTV", "会所", "棋牌", "俱乐部"
    ]
    
    # 优先寻找强后缀，并取最后出现的一个（以匹配最全的名称，如“集团股份有限公司”）
    best_match_end = -1
    
    for suffix in strong_suffixes:
        idx = s.rfind(suffix)
        if idx != -1:
            end_pos = idx + len(suffix)
            if end_pos > best_match_end:
                best_match_end = end_pos
    
    if best_match_end != -1:
        return s[:best_match_end].strip()
    
    # 如果没有强后缀，再尝试弱后缀
    for suffix in weak_suffixes:
        idx = s.rfind(suffix)
        if idx != -1:
            end_pos = idx + len(suffix)
            if end_pos > best_match_end:
                best_match_end = end_pos
                
    if best_match_end != -1:
        return s[:best_match_end].strip()
    
    return None

CUSTOM_MAPPINGS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "custom_mappings.json")

def load_custom_mappings():
    if not os.path.exists(CUSTOM_MAPPINGS_FILE):
        return {}
    try:
        with open(CUSTOM_MAPPINGS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}

def save_custom_mappings(mappings):
    try:
        with open(CUSTOM_MAPPINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(mappings, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"保存自定义映射失败: {e}")

def clean_group_company_entry(text: str) -> str:
    s = (text or "").strip()
    if not s:
        return ""
    s = re.sub(r'^\s*\d+\s*[.、]\s*', '', s)
    s = re.sub(r'^[（(]\s*保障中心\s*[）)]', '', s).strip()
    s = re.sub(r'[（(][^）)]*联系不上[^）)]*[）)]', '', s).strip()
    return s.strip()

def parse_groups(groups_file: str):
    groups = {}
    current_group = "未分组"
    last_line_was_empty = True # 初始视为有空行，以捕获第一行作为分类
    
    with open(groups_file, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line:
                last_line_was_empty = True
                continue
            
            # 如果上一行是空行，或者当前还没有确定分类，则这一行是分类名
            if last_line_was_empty:
                current_group = line
                if current_group not in groups:
                    groups[current_group] = []
                last_line_was_empty = False
            else:
                cleaned_company = clean_group_company_entry(line)
                if not cleaned_company:
                    last_line_was_empty = False
                    continue
                if re.search(r'\d', cleaned_company) and not is_company_line(cleaned_company):
                    last_line_was_empty = False
                    continue
                groups[current_group].append(cleaned_company)
                last_line_was_empty = False
                
    return groups

def get_company_to_group_map(groups: dict):
    company_to_group = {}
    for group_name, companies in groups.items():
        for company in companies:
            cleaned = clean_group_company_entry(company)
            norm = normalize_company(cleaned) or cleaned
            company_to_group[norm] = group_name
    return company_to_group

import csv
from datetime import datetime
from openpyxl import Workbook

from openpyxl.styles import Font, Alignment, PatternFill

import argparse

def process_single_file(data_file: str):
    """处理单个文件并返回通报数据列表"""
    try:
        with open(data_file, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
    except Exception as e:
        print(f"Error reading file {data_file}: {e}")
        return []

    # 跳过 HTTP 头
    start_idx = content.find("{")
    if start_idx == -1:
        # 尝试直接解析，可能没有 HTTP 头
        json_str = content
    else:
        json_str = content[start_idx:]
    
    try:
        data = json.loads(json_str)
    except json.JSONDecodeError:
        # 尝试修复可能的截断 JSON
        if "}" in json_str:
            last_brace = json_str.rfind("}")
            try:
                data = json.loads(json_str[:last_brace+1])
            except:
                print(f"Failed to recover JSON in {data_file}")
                return []
        else:
            print(f"No valid JSON in {data_file}")
            return []
    
    return data.get("notificationPigeonholeData", [])

def should_count_file(filename: str) -> bool:
    name = filename.strip()
    if not name:
        return False
    if name.startswith("~$"):
        return False
    lower = name.lower()
    if lower in {".ds_store", "thumbs.db", "desktop.ini"}:
        return False
    if lower.endswith((".tmp", ".part", ".crdownload")):
        return False
    return True

def is_notification_file(file_path: str) -> bool:
    name = os.path.basename(file_path).strip()
    if not name:
        return False
    if "通报" not in name:
        return False

    excluded_keywords = [
        "处置文件模板",
        "授权委托书",
        "责令整改通知书",
        "整改通知书",
        "处置报告",
        "复测报告",
        "整改报告",
        "营业执照",
        "身份证",
        "签字",
        "扫描",
        "回执",
    ]
    if any(k in name for k in excluded_keywords):
        return False

    lower = name.lower()
    allowed_exts = (".pdf", ".doc", ".docx", ".wps")
    if not lower.endswith(allowed_exts):
        return False
    return True

def looks_like_real_company(name: str) -> bool:
    s = (name or "").strip()
    if not s:
        return False
    strong_markers = (
        "股份有限公司",
        "有限责任公司",
        "有限公司",
        "责任有限公司",
        "集团公司",
        "集团",
        "公司",
        "制造厂",
        "工厂",
        "厂",
        "中心",
        "研究所",
        "研究院",
        "医院",
        "学校",
        "幼儿园",
        "托儿所",
        "商行",
        "事务所",
        "合作社",
        "农场",
        "经营部",
        "工作室",
        "委员会",
        "协会",
        "党支部",
        "联合会",
        "超市",
        "便利店",
        "饭店",
        "酒店",
        "宾馆",
        "旅馆",
    )
    if any(m in s for m in strong_markers):
        return True
    return False

def clean_candidate_company_text(text: str) -> str:
    s = (text or "").strip()
    if not s:
        return ""
    s = re.sub(r'[【\[][^】\]]*[】\]]', '', s)
    s = re.sub(r'[（(][^）)]*[）)]', '', s)
    s = re.sub(r'\d{4}[-_.年/]\d{1,2}[-_.月/]\d{1,2}日?', '', s)
    s = re.sub(r'\d{8}', '', s)
    s = re.sub(r'第?\s*\d+\s*期', '', s)
    s = re.sub(r'[\s_]+', '', s)
    s = s.replace("通报", "").replace("处置", "").replace("整改", "").replace("反馈", "").replace("已反馈", "")
    s = s.replace("留档", "").replace("运营中心", "").replace("网信办", "")
    s = s.strip("-—_·.，,;；:：")
    return s

def extract_company_from_file_path(file_path: str):
    name = os.path.splitext(os.path.basename(file_path))[0]
    candidates = [name]
    parent = os.path.dirname(file_path)
    for _ in range(3):
        if not parent:
            break
        base = os.path.basename(parent)
        if base and base not in candidates:
            candidates.append(base)
        next_parent = os.path.dirname(parent)
        if next_parent == parent:
            break
        parent = next_parent

    for cand in candidates:
        cleaned = clean_candidate_company_text(cand)
        if not cleaned:
            continue
        company = normalize_company(cleaned)
        if company and looks_like_real_company(company):
            return company
        if is_company_line(cleaned) and looks_like_real_company(cleaned):
            return cleaned
    return None

def detect_source_mode(input_path: str) -> str:
    def has_json_marker(path: str) -> bool:
        try:
            with open(path, "rb") as f:
                chunk = f.read(200_000)
            return b"notificationPigeonholeData" in chunk
        except Exception:
            return False

    if os.path.isfile(input_path):
        return "json" if has_json_marker(input_path) else "files"

    if os.path.isdir(input_path):
        for root, dirs, files in os.walk(input_path):
            for file in files:
                if not should_count_file(file):
                    continue
                file_path = os.path.join(root, file)
                return "json" if has_json_marker(file_path) else "files"
    return "json"

def is_geo_group_name(group_name: str) -> bool:
    s = (group_name or "").strip()
    if not s:
        return False
    if s in {"未分组", "联系不上"}:
        return False
    if any(k in s for k in ["镇", "街道", "街", "乡", "开发区", "园区", "新区"]):
        return True
    if s.endswith("区"):
        return True
    if s in {"潘火", "首南"}:
        return True
    return False

def run_statistics(input_path: str, groups_file: str, source: str = "auto"):
    # 1. 加载分类信息
    groups = parse_groups(groups_file)
    company_to_group = get_company_to_group_map(groups)
    
    if source == "auto":
        source = detect_source_mode(input_path)
        print(f"自动识别统计来源: {source}")
    
    company_counts = Counter()
    township_counts = Counter()
    township_to_companies = {} 
    company_to_township = {} 
    
    unmapped_companies = set()

    # 加载自定义映射
    custom_mappings = load_custom_mappings()
    
    # 预处理：应用自定义映射
    # 注意：这里需要修改 company_counts 的键
    original_companies = list(company_counts.keys())
    for company in original_companies:
        if company in custom_mappings:
            target_name = custom_mappings[company]
            count = company_counts.pop(company)
            company_counts[target_name] += count

    if source == "json":

        notifications = []
        if os.path.isfile(input_path):
            notifications = process_single_file(input_path)
        elif os.path.isdir(input_path):
            print(f"正在扫描文件夹: {input_path}")
            for root, dirs, files in os.walk(input_path):
                for file in files:
                    if file.endswith((".py", ".txt", ".csv", ".xlsx")):
                        continue
                    file_path = os.path.join(root, file)
                    print(f"  读取文件: {file}")
                    notifications.extend(process_single_file(file_path))
        else:
            print(f"错误: 路径不存在 {input_path}")
            return

        if not notifications:
            print("未发现有效的通报数据。")
            return

        print(f"共加载 {len(notifications)} 条通报数据。")

        for item in notifications:
            title = item.get("noticeTitle", "")
            if not title:
                continue
                
            company_name = normalize_company(title)
            if not company_name:
                continue
            
            company_counts[company_name] += 1

    elif source == "files":
        if not os.path.isdir(input_path):
            print("按文件数量统计模式需要输入一个文件夹路径。")
            return

        notification_files = 0
        for root, dirs, files in os.walk(input_path):
            for file in files:
                if not should_count_file(file):
                    continue
                file_path = os.path.join(root, file)
                if not is_notification_file(file_path):
                    continue
                notification_files += 1
                company_name = extract_company_from_file_path(file_path)
                if not company_name:
                    continue
                company_counts[company_name] += 1

        if notification_files == 0:
            print("未发现可统计的通报文件（文件名需包含“通报”）。")
            return
        if not company_counts:
            print("发现通报文件但未能从文件名/路径提取企业名称。")
            return

        print(f"共扫描到 {notification_files} 份通报文件，识别到 {len(company_counts)} 家企业。")
    else:
        print(f"错误: 未知统计来源 {source}")
        return

    for company_name, count in company_counts.items():
        township = company_to_group.get(company_name)
        if not township:
            unmapped_companies.add(company_name)
            continue

        township_counts[township] += count
        company_to_township[company_name] = township
        
        if township not in township_to_companies:
            township_to_companies[township] = Counter()
        township_to_companies[township][company_name] += count

    # 4. 判断输出逻辑
    if unmapped_companies:
        print("\n" + "!"*60)
        print("发现未识别分类的企业，请更新 1.txt 后再次运行脚本")
        print("!"*60)
        print("\n[待查询企业名单]")
        unmapped_list = sorted([(c, company_counts[c]) for c in unmapped_companies], key=lambda x: x[1], reverse=True)
        for i, (comp, count) in enumerate(unmapped_list, 1):
            print(f"{i}. {comp}")
        print("\n提示: 请将上述企业名称复制并分配到 1.txt 中的对应镇街下方。")

    # 如果没有未识别的企业，则输出表格
    print("\n" + "="*60)
    print("所有企业已识别镇街，正在生成统计报表...")
    print("="*60)

    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.dirname(os.path.abspath(__file__))
    
    csv_file = f"statistics_report_{timestamp}.csv"
    xlsx_file = f"statistics_report_{timestamp}.xlsx"
    csv_path = os.path.join(output_dir, csv_file)
    xlsx_path = os.path.join(output_dir, xlsx_file)
    
    try:
        township_order = [
            "瞻岐镇",
            "咸祥镇",
            "东吴镇",
            "塘溪镇",
            "五乡镇",
            "邱隘镇",
            "云龙镇",
            "横溪镇",
            "姜山镇",
            "东钱湖镇",
            "潘火街道",
            "福明街道",
            "东柳街道",
            "中河街道",
            "东郊街道",
            "下应街道",
            "明楼街道",
            "百丈街道",
            "东胜街道",
            "白鹤街道",
            "首南街道",
            "钟公庙街道",
            "南部商务区",
            "经济开发区",
        ]
        excluded_groups = {"未分组", "联系不上"}
        group_order_from_file = [g for g in groups.keys() if g not in excluded_groups]
        selected_set = set()
        sorted_groups = []
        for g in township_order:
            if g in group_order_from_file and g not in selected_set:
                sorted_groups.append(g)
                selected_set.add(g)
        for g in group_order_from_file:
            if g not in selected_set:
                sorted_groups.append(g)
                selected_set.add(g)

        # 写入 CSV
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["镇街名称", "合计"])
            for township in sorted_groups:
                total_count = township_counts.get(township, 0)
                comp_list = township_to_companies.get(township, Counter())
                company_count = len(comp_list)
                writer.writerow([township, f"企业 {company_count} 家，通报 {total_count} 次"])
        
        # 写入 XLSX
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("无法创建 Excel 工作表")
            
        ws.title = "通报统计报表"
        
        # 表头
        headers = ["镇街名称", "合计"]
        ws.append(headers)
        
        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        township_font = Font(bold=True, color="000000")
        township_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_aligned
            
        # 写入主统计数据
        current_data_row = 2
        for township in sorted_groups:
            total_count = township_counts.get(township, 0)
            comp_list = township_to_companies.get(township, Counter())
            company_count = len(comp_list)
            
            # 写入镇街汇总行
            ws.cell(row=current_data_row, column=1, value=township)
            ws.cell(row=current_data_row, column=2, value=f"企业 {company_count} 家，通报 {total_count} 次")
            
            # 设置汇总行样式
            for col in range(1, 3):
                cell = ws.cell(row=current_data_row, column=col)
                cell.font = township_font
                cell.fill = township_fill
                cell.alignment = left_aligned
            
            current_data_row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 28
        
        wb.save(xlsx_path)
        
        print(f"\n✅ 报表生成成功:")
        print(f"   CSV:  {csv_path}")
        print(f"   XLSX: {xlsx_path}")
        
        print("\n[全区 Top 10 企业预览]")
        top_10 = company_counts.most_common(10)
        for i, (company, count) in enumerate(top_10, 1):
            township = company_to_township.get(company)
            print(f"{i:2}. {company:30} ({township}) | {count} 次")
            
    except Exception as e:
        print(f"生成报表失败: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="网信办通报数据分类统计脚本")
    parser.add_argument("input", nargs="?", default=r"c:\Users\lan1o\Desktop\wow\1", help="数据文件路径或包含文件的文件夹路径")
    parser.add_argument("--groups", default=r"c:\Users\lan1o\Desktop\wow\1.txt", help="分组定义文件路径")
    parser.add_argument("--source", choices=["auto", "json", "files"], default="auto", help="统计来源：json=接口留档；files=按文件数量统计；auto=自动识别")
    
    args = parser.parse_args()
    run_statistics(args.input, args.groups, source=args.source)
